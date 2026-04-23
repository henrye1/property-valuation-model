from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from valuation_engine import calculate
from valuation_engine.excel.render import render_workbook
from valuation_engine.models import ParkingLine, TenantLine, ValuationInput


def _sample_input() -> ValuationInput:
    return ValuationInput(
        valuation_date=date(2026, 4, 1),
        tenants=[TenantLine(
            description="Offices",
            rentable_area_m2=Decimal("182"),
            rent_per_m2_pm=Decimal("85"),
            annual_escalation_pct=Decimal("0"),
        )],
        parking=[ParkingLine(bay_type="open", bays=2, rate_per_bay_pm=Decimal("200"))],
        monthly_operating_expenses=Decimal("2910.17"),
        vacancy_allowance_pct=Decimal("0"),
        cap_rate=Decimal("0.11"),
    )


def test_render_writes_canonical_labels(tmp_path: Path):
    inputs = _sample_input()
    result = calculate(inputs)
    out = tmp_path / "out.xlsx"
    render_workbook(out, building_name="Test Building", inputs=inputs, result=result)

    wb = load_workbook(out, data_only=False)
    ws = wb.active
    labels = {
        (ws.cell(row=r, column=1).value or "").strip().lower()
        for r in range(1, ws.max_row + 1)
    }
    for needle in (
        "building name :",
        "date",
        "sub total",
        "parking",
        "gross monthly income",
        "gross annual income",
        "operating expenses",
        "vacancy allowance",
        "monthly net income",
        "annual net income",
        "capitalised @",
        "open market assessment",
    ):
        assert any(needle in label for label in labels), f"Missing label: {needle}"
