"""Write a valuation snapshot to an Excel workbook in the canonical layout.

Formulas are preserved (not just values) so a recipient can audit the math
in Excel.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from valuation_engine.models import ValuationInput, ValuationResult


def render_workbook(
    out_path: Path,
    *,
    building_name: str,
    inputs: ValuationInput,
    result: ValuationResult,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"

    # --- Header ---
    ws.cell(row=1, column=1, value="Building name : ")
    ws.cell(row=1, column=2, value=building_name)
    ws.cell(row=2, column=1, value="Date")
    ws.cell(row=2, column=5, value=inputs.valuation_date)

    # --- Tenant header ---
    ws.cell(row=3, column=1, value="Tenant")
    ws.cell(row=3, column=2, value="Description")
    ws.cell(row=3, column=3, value="Annual escalation")
    ws.cell(row=3, column=4, value="Lease period")
    ws.cell(row=3, column=5, value="Lease expiry date")
    ws.cell(row=3, column=6, value="Rentable area")
    ws.cell(row=3, column=7, value="Gross/Net rent per month (R/m2/pm)")
    ws.cell(row=3, column=8, value="Monthly operating expenses (R/m2/pm)")
    ws.cell(row=3, column=9, value="Gross monthly rent")

    tenant_start = 4
    for i, t in enumerate(inputs.tenants):
        r = tenant_start + i
        ws.cell(row=r, column=2, value=t.description)
        if t.annual_escalation_pct:
            ws.cell(row=r, column=3, value=float(t.annual_escalation_pct))
        ws.cell(row=r, column=4, value=t.lease_period_text or "")
        ws.cell(row=r, column=5, value=t.lease_expiry_date)
        ws.cell(row=r, column=6, value=float(t.rentable_area_m2))
        ws.cell(row=r, column=7, value=float(t.rent_per_m2_pm))
        ws.cell(row=r, column=9, value=f"=G{r}*F{r}")
    tenant_end = tenant_start + len(inputs.tenants) - 1
    tenant_subtotal_row = tenant_end + 1
    ws.cell(row=tenant_subtotal_row, column=1, value="Sub total")
    ws.cell(row=tenant_subtotal_row, column=6, value=f"=SUM(F{tenant_start}:F{tenant_end})")
    ws.cell(row=tenant_subtotal_row, column=9, value=f"=SUM(I{tenant_start}:I{tenant_end})")

    # --- Parking ---
    parking_label_row = tenant_subtotal_row + 2
    ws.cell(row=parking_label_row, column=1, value="Parking")
    ws.cell(row=parking_label_row, column=6, value="No. bays")
    ws.cell(row=parking_label_row, column=7, value="R/bay")

    parking_start = parking_label_row + 1
    for i, p in enumerate(inputs.parking):
        r = parking_start + i
        ws.cell(row=r, column=5, value=p.bay_type)
        ws.cell(row=r, column=6, value=p.bays)
        ws.cell(row=r, column=7, value=float(p.rate_per_bay_pm))
        ws.cell(row=r, column=9, value=f"=G{r}*F{r}")
    parking_end = parking_start + len(inputs.parking) - 1 if inputs.parking else parking_start
    parking_subtotal_row = parking_end + 1
    ws.cell(row=parking_subtotal_row, column=1, value="Sub total")
    if inputs.parking:
        ws.cell(row=parking_subtotal_row, column=6, value=f"=SUM(F{parking_start}:F{parking_end})")
        ws.cell(row=parking_subtotal_row, column=9, value=f"=SUM(I{parking_start}:I{parking_end})")
    else:
        ws.cell(row=parking_subtotal_row, column=9, value=0)

    # --- Income aggregation ---
    gmi_row = parking_subtotal_row + 1
    ws.cell(row=gmi_row, column=1, value="Gross monthly income")
    ws.cell(row=gmi_row, column=9, value=f"=I{parking_subtotal_row}+I{tenant_subtotal_row}")
    gai_row = gmi_row + 1
    ws.cell(row=gai_row, column=1, value="Gross annual income")
    ws.cell(row=gai_row, column=9, value=f"=I{gmi_row}*12")

    # --- Operating expenses ---
    opex_label_row = gai_row + 1
    ws.cell(row=opex_label_row, column=1, value="Operating expenses")
    ws.cell(row=opex_label_row, column=5, value="Monthly")
    ws.cell(row=opex_label_row, column=6, value="Annual")
    ws.cell(row=opex_label_row, column=7, value="R/m2/pm")
    opex_subtotal_row = opex_label_row + 1
    ws.cell(row=opex_subtotal_row, column=1, value="Sub total")
    ws.cell(row=opex_subtotal_row, column=5, value=float(inputs.monthly_operating_expenses))
    ws.cell(row=opex_subtotal_row, column=6, value=f"=E{opex_subtotal_row}*12")
    ws.cell(row=opex_subtotal_row, column=7, value=f"=E{opex_subtotal_row}/F{tenant_subtotal_row}")
    ws.cell(row=opex_subtotal_row, column=8, value=f"=F{opex_subtotal_row}/I{gai_row}")

    vacancy_row = opex_subtotal_row + 1
    ws.cell(row=vacancy_row, column=1, value="Vacancy allowance")
    ws.cell(row=vacancy_row, column=8, value=float(inputs.vacancy_allowance_pct))
    ws.cell(row=vacancy_row, column=9, value=f"=I{gai_row}*-H{vacancy_row}")

    mni_row = vacancy_row + 1
    ws.cell(row=mni_row, column=1, value="Monthly net income")
    ani_row = mni_row + 1
    ws.cell(row=ani_row, column=1, value="Annual net income")
    ws.cell(row=ani_row, column=9, value=f"=I{gai_row}-F{opex_subtotal_row}+I{vacancy_row}")
    ws.cell(row=mni_row, column=9, value=f"=I{ani_row}/12")

    cap_row = ani_row + 1
    ws.cell(row=cap_row, column=1, value="Capitalised @")
    ws.cell(row=cap_row, column=8, value=float(inputs.cap_rate))
    ws.cell(row=cap_row, column=9, value=f"=I{ani_row}/H{cap_row}")

    omv_row = cap_row + 1
    ws.cell(row=omv_row, column=1, value="Open market assessment")
    if inputs.rounding == "nearest_10000":
        rounding_arg = -4
    elif inputs.rounding == "nearest_1000":
        rounding_arg = -3
    else:
        rounding_arg = 0
    ws.cell(row=omv_row, column=9, value=f"=ROUND(I{cap_row},{rounding_arg})")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
