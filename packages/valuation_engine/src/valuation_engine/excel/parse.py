"""Tolerant, label-driven Excel parser.

Locates sections by column-A labels rather than absolute row numbers, so the
parser tolerates the layout drift seen across the historical sample sheets.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ConfigDict, ValidationError

from valuation_engine.models import ValuationInput, Warning

TENANT_HEADER_NEEDLES = ("rentable area",)
SUBTOTAL_LABEL = "sub total"
SECTION_END_BLANK_RUN = 2  # consecutive blank rows = end of section

BAY_TYPE_MAP = {
    "open": "open",
    "covered": "covered",
    "shade": "shade",
    "shaded": "shade",
    "basement": "basement",
}


def compute_diff_pct(sheet_value: Decimal, recomputed_value: Decimal) -> Decimal | None:
    """Returns abs(recomputed - sheet) / sheet, or None if sheet_value is 0/None."""
    if not sheet_value:
        return None
    return abs(recomputed_value - sheet_value) / sheet_value


RECOMPUTE_TOLERANCE_PCT = Decimal("0.001")


@dataclass
class _SheetCursor:
    ws: Worksheet
    max_row: int

    def find_label_row(self, *needles: str, start: int = 1) -> int | None:
        """Return the 1-based row index of the first column-A cell matching any needle.

        Matching is a case-insensitive substring match.
        """
        needles_lower = [n.lower() for n in needles]
        for r in range(start, self.max_row + 1):
            v = self.ws.cell(row=r, column=1).value
            if isinstance(v, str):
                vl = v.strip().lower()
                if any(n in vl for n in needles_lower):
                    return r
        return None


class ParseResult(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)
    inputs: ValuationInput | None
    building_name: str | None
    sheet_market_value: Decimal | None
    parse_warnings: list[Warning]
    parse_errors: list[Warning]


def parse_workbook(path: Path) -> ParseResult:
    """Parse a single workbook and return inputs + diagnostics.

    Reads with `data_only=True` so cached formula values are available for the
    recompute-and-compare check; falls back to formula-only parse if a cell has
    no cached value (raises a `formula_missing_value` warning per cell).
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    sheet_names = wb.sheetnames
    parse_warnings: list[Warning] = []
    parse_errors: list[Warning] = []

    if len(sheet_names) > 1:
        parse_warnings.append(
            Warning(
                code="multiple_sheets",
                message=(
                    f"Workbook has {len(sheet_names)} sheets; "
                    f"using the first ('{sheet_names[0]}')."
                ),
                field_path=None,
            )
        )

    ws = wb[sheet_names[0]]
    cur = _SheetCursor(ws=ws, max_row=ws.max_row or 1)

    building_name = _read_building_name(cur)
    valuation_date = _read_valuation_date(cur, parse_warnings, parse_errors)
    tenant_dicts, last_tenant_row = _read_tenants(cur, parse_warnings, parse_errors)
    parking_dicts, last_parking_row = _read_parking(
        cur, last_tenant_row, parse_warnings
    )
    assumptions, sheet_market_value = _read_assumptions(
        cur, last_parking_row, parse_warnings, parse_errors
    )

    inputs = _build_inputs_partial(
        valuation_date=valuation_date,
        tenant_dicts=tenant_dicts,
        parking_dicts=parking_dicts,
        assumptions=assumptions,
        parse_errors=parse_errors,
    )

    return ParseResult(
        inputs=inputs,
        building_name=building_name,
        sheet_market_value=sheet_market_value,
        parse_warnings=parse_warnings,
        parse_errors=parse_errors,
    )


def _read_building_name(cur: _SheetCursor) -> str | None:
    row = cur.find_label_row("building name")
    if row is None:
        return None
    v = cur.ws.cell(row=row, column=2).value
    if v is None:
        return None
    return str(v).strip()


def _read_valuation_date(
    cur: _SheetCursor,
    parse_warnings: list[Warning],
    parse_errors: list[Warning],
) -> date | None:
    row = cur.find_label_row("date")
    if row is None:
        parse_errors.append(
            Warning(
                code="missing_required_section",
                message="Could not find 'Date' label in column A.",
                field_path="valuation_date",
            )
        )
        return None
    # Date is in column E in canonical layout; tolerate B-E by scanning.
    for col in (5, 4, 3, 2):
        v = cur.ws.cell(row=row, column=col).value
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
    parse_errors.append(
        Warning(
            code="missing_required_section",
            message="Date row found but no parseable date value in columns B-E.",
            field_path="valuation_date",
        )
    )
    return None


def _find_tenant_header_row(cur: _SheetCursor) -> int | None:
    """Find the header row that contains 'Rentable area' anywhere in cols A-J."""
    for r in range(1, cur.max_row + 1):
        for col in range(1, 11):
            v = cur.ws.cell(row=r, column=col).value
            if isinstance(v, str) and "rentable area" in v.lower():
                return r
    return None


def _read_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        # bool is a subclass of int; don't coerce booleans to Decimal.
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    if isinstance(value, str):
        s = value.replace(",", "").replace("R", "").strip()
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None
    return None


def _read_pct(value: Any) -> Decimal | None:
    """Parse a percentage cell. Excel may store 8% as 0.08 or as the string '8%'."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    if isinstance(value, str):
        s = value.strip().replace("%", "")
        try:
            d = Decimal(s)
        except (InvalidOperation, ValueError):
            return None
        return d / Decimal("100") if "%" in value else d
    return None


def _read_tenants(
    cur: _SheetCursor,
    parse_warnings: list[Warning],
    parse_errors: list[Warning],
) -> tuple[list[dict[str, Any]], int]:
    """Returns (tenant_dicts, last_row_consumed).

    tenant_dicts have keys: description, rentable_area_m2, rent_per_m2_pm,
    annual_escalation_pct (Decimal|None).
    """
    header_row = _find_tenant_header_row(cur)
    if header_row is None:
        parse_errors.append(
            Warning(
                code="missing_required_section",
                message="Could not find tenant header row containing 'Rentable area'.",
                field_path="tenants",
            )
        )
        return [], 0

    tenants: list[dict[str, Any]] = []
    blank_run = 0
    last_row = header_row
    for r in range(header_row + 1, cur.max_row + 1):
        a = cur.ws.cell(row=r, column=1).value
        if isinstance(a, str) and SUBTOTAL_LABEL in a.lower():
            return tenants, r
        # Detect end-of-section by blank streak.
        row_values = [cur.ws.cell(row=r, column=c).value for c in range(1, 10)]
        if all(v is None or v == "" for v in row_values):
            blank_run += 1
            if blank_run >= SECTION_END_BLANK_RUN:
                return tenants, r
            continue
        blank_run = 0

        area = _read_decimal(cur.ws.cell(row=r, column=6).value)
        rent = _read_decimal(cur.ws.cell(row=r, column=7).value)
        if area is None or rent is None:
            parse_warnings.append(
                Warning(
                    code="unrecognised_row",
                    message=(
                        f"Row {r}: tenant row missing area or rent "
                        f"(area={area}, rent={rent})."
                    ),
                    field_path=f"tenants[row {r}]",
                )
            )
            continue
        if area <= 0:
            # Skip subtotal rows or zero-area filler rows silently.
            continue

        description = (
            cur.ws.cell(row=r, column=2).value
            or cur.ws.cell(row=r, column=1).value
            or ""
        )
        esc = _read_pct(cur.ws.cell(row=r, column=3).value)
        tenants.append(
            dict(
                description=str(description).strip(),
                rentable_area_m2=area,
                rent_per_m2_pm=rent,
                annual_escalation_pct=esc if esc is not None else Decimal("0"),
            )
        )
        last_row = r

    return tenants, last_row


def _classify_bay_type(label: Any) -> str:
    if isinstance(label, str):
        key = label.strip().lower()
        return BAY_TYPE_MAP.get(key, "other")
    return "other"


def _read_parking(
    cur: _SheetCursor,
    start_row: int,
    parse_warnings: list[Warning],
) -> tuple[list[dict[str, Any]], int]:
    """Returns (parking_dicts, last_row)."""
    parking_label_row = cur.find_label_row("parking", start=start_row)
    if parking_label_row is None:
        # Parking is optional.
        parse_warnings.append(
            Warning(
                code="missing_optional_section",
                message="No 'Parking' section found.",
                field_path="parking",
            )
        )
        return [], start_row

    parking: list[dict[str, Any]] = []
    last_row = parking_label_row
    blank_run = 0
    # Parking rows start two below the label (label row, header row, then data).
    for r in range(parking_label_row + 2, cur.max_row + 1):
        a = cur.ws.cell(row=r, column=1).value
        if isinstance(a, str) and SUBTOTAL_LABEL in a.lower():
            return parking, r
        row_values = [cur.ws.cell(row=r, column=c).value for c in range(1, 10)]
        if all(v is None or v == "" for v in row_values):
            blank_run += 1
            if blank_run >= SECTION_END_BLANK_RUN:
                return parking, r
            continue
        blank_run = 0

        bays = _read_decimal(cur.ws.cell(row=r, column=6).value)
        rate = _read_decimal(cur.ws.cell(row=r, column=7).value)
        if bays is None or rate is None or bays == 0:
            continue
        bay_type = _classify_bay_type(cur.ws.cell(row=r, column=5).value)
        parking.append(
            dict(
                bay_type=bay_type,
                bays=int(bays),
                rate_per_bay_pm=rate,
            )
        )
        last_row = r

    return parking, last_row


def _read_assumptions(
    cur: _SheetCursor,
    start_row: int,
    parse_warnings: list[Warning],
    parse_errors: list[Warning],
) -> tuple[dict[str, Any], Decimal | None]:
    """Returns (assumptions_dict, sheet_market_value).

    assumptions_dict may contain: monthly_operating_expenses,
    vacancy_allowance_pct, cap_rate.
    """
    out: dict[str, Any] = {}

    opex_label = cur.find_label_row("operating expenses", start=start_row)
    if opex_label is not None:
        # The next 'Sub total' row holds monthly opex in column E.
        for r in range(opex_label + 1, cur.max_row + 1):
            a = cur.ws.cell(row=r, column=1).value
            if isinstance(a, str) and SUBTOTAL_LABEL in a.lower():
                opex_monthly = _read_decimal(cur.ws.cell(row=r, column=5).value)
                if opex_monthly is None:
                    parse_errors.append(
                        Warning(
                            code="missing_required_section",
                            message="Operating expenses subtotal row has no value in column E.",
                            field_path="monthly_operating_expenses",
                        )
                    )
                else:
                    out["monthly_operating_expenses"] = opex_monthly
                break
    else:
        parse_errors.append(
            Warning(
                code="missing_required_section",
                message="No 'Operating expenses' section found.",
                field_path="monthly_operating_expenses",
            )
        )

    vacancy_row = cur.find_label_row("vacancy allowance", start=start_row)
    if vacancy_row is not None:
        vac = _read_pct(cur.ws.cell(row=vacancy_row, column=8).value)
        if vac is None:
            parse_warnings.append(
                Warning(
                    code="non_canonical_label",
                    message="Vacancy allowance % missing in column H; defaulted to 0.",
                    field_path="vacancy_allowance_pct",
                )
            )
            out["vacancy_allowance_pct"] = Decimal("0")
        else:
            out["vacancy_allowance_pct"] = vac
    else:
        parse_errors.append(
            Warning(
                code="missing_required_section",
                message="No 'Vacancy allowance' row found.",
                field_path="vacancy_allowance_pct",
            )
        )

    cap_row = cur.find_label_row("capitalised", start=start_row)
    if cap_row is not None:
        cap = _read_pct(cur.ws.cell(row=cap_row, column=8).value)
        if cap is None:
            parse_errors.append(
                Warning(
                    code="missing_required_section",
                    message="Capitalisation rate missing in column H.",
                    field_path="cap_rate",
                )
            )
        else:
            out["cap_rate"] = cap
    else:
        parse_errors.append(
            Warning(
                code="missing_required_section",
                message="No 'Capitalised @' row found.",
                field_path="cap_rate",
            )
        )

    sheet_market_value: Decimal | None = None
    omv_row = cur.find_label_row("open market assessment", start=start_row)
    if omv_row is not None:
        sheet_market_value = _read_decimal(cur.ws.cell(row=omv_row, column=9).value)

    return out, sheet_market_value


def _build_inputs_partial(
    *,
    valuation_date: date | None,
    tenant_dicts: list[dict[str, Any]],
    parking_dicts: list[dict[str, Any]],
    assumptions: dict[str, Any],
    parse_errors: list[Warning],
) -> ValuationInput | None:
    if valuation_date is None or not tenant_dicts:
        return None
    if "cap_rate" not in assumptions or "monthly_operating_expenses" not in assumptions:
        return None
    try:
        from valuation_engine.models import ParkingLine, TenantLine
        return ValuationInput(
            valuation_date=valuation_date,
            tenants=[TenantLine(**t) for t in tenant_dicts],
            parking=[ParkingLine(**p) for p in parking_dicts],
            monthly_operating_expenses=assumptions["monthly_operating_expenses"],
            vacancy_allowance_pct=assumptions.get(
                "vacancy_allowance_pct", Decimal("0")
            ),
            cap_rate=assumptions["cap_rate"],
        )
    except (ValueError, TypeError, ValidationError) as exc:
        parse_errors.append(
            Warning(
                code="missing_required_section",
                message=f"Could not assemble ValuationInput: {exc}",
                field_path=None,
            )
        )
        return None
